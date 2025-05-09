
from datetime import date, datetime
from .RoomData import RoomData
import openpyxl as xl
from openpyxl.styles import *

from ..Logging import LoggingManager as logging



class DailyStatistics():
    """
    ### 일간 출석부 데이터

    특정 일자의 RoomData를 기반으로 학번, 이름, 입실 시간, 퇴실 시간, 좌석, 이동 횟수를 확인해 통계를 만드는 클래스.
    """

    
    Statistics: list[list[str]]
    # [<StudentID: str>, <Name: str>, <LastChkIn: time>, <LastChkOut: time>, <LastSeat: str>, <MoveCount: int>]

    Empty: bool # 비어있음 여부
    TotalMove: int # 전체 누적 이동 수



    def __init__(self, Date: date = None):
        """
        #### 매개변수:
        * **Date:** 내보내고자 할 일자
        """

        # 날짜 설정
        if Date != None:
            self.DATE = Date
        else:
            self.DATE = RoomData.DATA_DATE

        # 시작 로그
        logging.info(self.DATE.strftime('일간 출석부 기록을 불러오는 중: %Y/%m/%d'))

        # RoomData로부터 학생 데이터 취합
        Students = {}

        RD: RoomData = RoomData.Load(self.DATE)

        if RD == None:
            self.Empty = True
            return
        else:
            self.Empty = False

        self.TotalMove = 0

        for log in RD.Logs: # 기록 순회

            # Students 딕셔너리에 초기 데이터 기입
            id = log['ID']
            if id not in Students:
                Students[id] = [id, log['Name'], None, None, None, 0]

            # 입실 기록인 경우
            if log['Action'] == 'ChkIn':

                Students[id][4] = log['Seat']
                timestamp = datetime.strptime(log['TimeStamp'], '%Y%m%d%H%M%S.%f').time()

                if Students[id][2] == None or Students[id][2] > timestamp:
                    Students[id][2] = timestamp
                
            # 퇴실 기록인 경우
            elif log['Action'] == 'ChkOut':
                
                timestamp = datetime.strptime(log['TimeStamp'], '%Y%m%d%H%M%S.%f').time()

                if Students[id][3] == None or Students[id][3] < timestamp:
                    Students[id][3] = timestamp

            # 좌석 이동 기록인 경우
            elif log['Action'] == 'Move':
                Students[id][4] = log['Seat']
                Students[id][5] += 1
                self.TotalMove += 1

        # 읽을 수 있는 문자열로 변환
        self.Statistics = []

        for student in Students.values():
            self.Statistics.append([
                student[0], student[1], 
                DailyStatistics.format_time(student[2]), 
                DailyStatistics.format_time(student[3]), 
                student[4], student[5]
            ])



    @staticmethod
    def format_timeStr(dtStr: str) -> str:
        if str(dtStr).strip().lower() in ('none', 'null', ''):
            return '?'

        dt = datetime.strptime(dtStr, '%Y%m%d%H%M%S.%f')

        # 오전/오후 결정
        period = "오전" if dt.hour < 12 else "오후"

        # 12시간제로 변환
        hour = dt.hour % 12
        if hour == 0:
            hour = 12

        # 초는 소수 첫째자리까지만
        seconds = f"{dt.second if dt.second > 9 else '0' + str(dt.second)}.{int(dt.microsecond / 100_000)}"

        # 최종 포맷
        return f"{period} {hour:02d}:{dt.minute:02d}:{seconds}"
    


    @staticmethod
    def cell(sheet, row, column, value):
        sheet.cell(row=row + 3, column=column).value = value
        sheet.cell(row=row + 3, column=column).border = Border(left=Side(style='thin'),
                                                          right=Side(style='thin'),
                                                          top=Side(style='thin'),
                                                          bottom=Side(style='thin'))
        sheet.cell(row=row + 3, column=column).alignment = Alignment(vertical='center', horizontal='center')

    
    def Write(self) -> str:
        """ 통계를 작성하고 파일로 내보냄. 오류가 발생하면 기록을 출력하고 ErrorDialog를 표시함. """

        try:
            logging.info(self.DATE.strftime('일간 출석부 통계 작성 시작: %Y/%m/%d'))
            path = self._Write()
            logging.info('일간 출석부 통계 작성 완료: ' + path)
            return path
        except Exception as e:
            logging.error('일간 출석부 통계 작성 중 오류가 발생하였습니다.', e, False, True)
            return None


    def _Write(self) -> str:
        
        from ..Info import ChairyInfo as CI

        wb: xl.Workbook = xl.load_workbook(RoomData.DIRECTORY + "/ChairyApp/sheets/daily.xlsx")

        ws = wb.active

        for row, value in enumerate(self.Statistics, start=1):  # 예를 들면 2번째 행부터 시작
            DailyStatistics.cell(ws, row, 1, row)
            DailyStatistics.cell(ws, row, 2, value[0])
            DailyStatistics.cell(ws, row, 3, value[1])
            DailyStatistics.cell(ws, row, 4, value[2])
            DailyStatistics.cell(ws, row, 5, value[3])
            DailyStatistics.cell(ws, row, 6, value[4])
            DailyStatistics.cell(ws, row, 7, value[5])

        ws.cell(row=3, column=9).value = RoomData.DATA_DATE
        ws.cell(row=4, column=9).value = len(RoomData.DATA.Arrangement)

        FirstChkInTime: datetime = None
        FirstChkInStudent: str = '?'
        FirstChkOutTime: datetime = None
        FirstChkOutStudent: str = '?'

        LastChkInTime: datetime = None
        LastChkInStudent: str = '?'
        LastChkOutTime: datetime = None
        LastChkOutStudent: str = '?'

        for data in RoomData.DATA.Logs:
            if data['Action'] == 'ChkIn':
                time = datetime.strptime(data['TimeStamp'], "%Y%m%d%H%M%S.%f")
                if FirstChkInTime is None or FirstChkInTime > time:
                    FirstChkInTime = time
                    FirstChkInStudent = data['ID'] + " " + data['Name']

                if LastChkInTime is None or LastChkInTime < time:
                    LastChkInTime = time
                    LastChkInStudent = data['ID'] + " " + data['Name']

            elif data['Action'] == 'ChkOut':
                time = datetime.strptime(data['TimeStamp'], "%Y%m%d%H%M%S.%f")
                if FirstChkOutTime is None or FirstChkOutTime > time:
                    FirstChkOutTime = time
                    FirstChkOutStudent = data['ID'] + " " + data['Name']

                if LastChkOutTime is None or LastChkOutTime < time:
                    LastChkOutTime = time
                    LastChkOutStudent = data['ID'] + " " + data['Name']


        ws.cell(row=5, column=9).value = DailyStatistics.format_time(FirstChkInTime)
        ws.cell(row=6, column=9).value = DailyStatistics.format_time(FirstChkOutTime)
        ws.cell(row=7, column=9).value = FirstChkInStudent
        ws.cell(row=8, column=9).value = FirstChkOutStudent

        ws.cell(row=9, column=9).value = DailyStatistics.format_time(LastChkInTime)
        ws.cell(row=10, column=9).value = DailyStatistics.format_time(LastChkOutTime)
        ws.cell(row=11, column=9).value = LastChkInStudent
        ws.cell(row=12, column=9).value = LastChkOutStudent

        ws.cell(row=13, column=9).value = len(RoomData.DATA.UserNames)
        ws.cell(row=14, column=9).value = self.TotalMove

        ws.cell(row=15, column=9).value = DailyStatistics.format_time(RoomData.DATA.Begin)
        ws.cell(row=16, column=9).value = DailyStatistics.format_time(RoomData.DATA.End)

        if RoomData.TodayReservedSeat():
            ws.cell(row=17, column=9).value = 'O'
        else:
            ws.cell(row=17, column=9).value = 'X'

        ws.cell(row=18, column=9).value = CI.Version

        for row in range(4, ws.max_row):
            ws.row_dimensions[row].height = 20

        import os
        if not os.path.exists(RoomData.DIRECTORY + '/Statistics/'):
            os.makedirs(RoomData.DIRECTORY + '/Statistics/')

        finalFile = RoomData.DATA_DATE.strftime('일간 출석부-%Y%m%d-') + datetime.now().strftime('%Y%m%d%H%M%S') + '.xlsx'
        wb.save(RoomData.DIRECTORY + '/Statistics/' + finalFile)
        
        return finalFile



    @staticmethod
    def format_time(dt: datetime) -> str:
        if dt == None:
            return '?'
        
        # 오전/오후 결정
        period = "오전" if dt.hour < 12 else "오후"

        # 12시간제로 변환
        hour = dt.hour % 12
        if hour == 0:
            hour = 12

        # 초는 소수 첫째자리까지만
        seconds = f"{dt.second if dt.second > 9 else '0' + str(dt.second)}.{int(dt.microsecond / 100_000)}"

        # 최종 포맷
        return f"{period} {hour:02d}:{dt.minute:02d}:{seconds}"