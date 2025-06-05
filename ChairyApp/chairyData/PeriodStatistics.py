
from os import listdir
from os.path import exists
from datetime import date, datetime, timedelta
from .RoomData import RoomData
import openpyxl as xl
from openpyxl.styles import *
from dataclasses import dataclass

from ..Logging import LoggingManager as logging



dataclass(slots=True)
class PeriodStatistics():
    """
    ### 교시별 출석부 데이터

    학생별 입,퇴실 시간을 분석하여 Configuration에 입력된 자습 시간에 따라 기호로 해당 교시에서의 자습 참여 여부를 통계로 만드는 클래스.
    """


    Year    : int
    Month   : int

    FileDates: list[date] # RoomData가 존재하는 일자

    Statistics: dict
    # {
    #   '<StudentID>' : [ [<SelfStudyTime1>, <SelfStudyTime2>, <SelfStudyTime3>], ... ]
    # }

    Empty: bool # 비어있음 여부



    def __init__(self, year: int, month: int):
        """
        - **year:** 내보내고자 하는 년도
        - **month:** 내보내고자 하는 달
        """

        # 날짜 설정
        self.Year  = year
        self.Month = month

        # 파일 이름 가져오고 날짜 추출
        if not exists(RoomData.DIRECTORY + date(year, month, 1).strftime('/RoomData/%Y%m/')):
            self.Empty = True
            return

        self.FileDates = []

        for fn in listdir(RoomData.DIRECTORY + date(year, month, 1).strftime('/RoomData/%Y%m/')):
            if fn.endswith('.json'):
                name = fn.split('.json')[0]
                try:
                    self.FileDates.append(datetime.strptime(name, '%Y%m%d').date())
                except Exception:
                    continue

        if len(self.FileDates) < 1:
            self.Empty = True
            return
        else:
            self.Empty = False

        # 시작 로그
        logging.info(f'교시별 출석부 기록을 불러오는 중: {year}년 {month}월')

        # _DATA_DATES를 기본 포함해서 딕셔너리 구성
        self.Statistics = {
            '_DATA_DATES' : [dt.strftime('%Y%m%d') for dt in self.FileDates],
            '_NAMES' : {}
        }

        # 파일 별로 RoomData 불러오고 입,퇴실 시간 취합
        times = {}

        for index, dt in enumerate(self.FileDates):

            roomdata = RoomData.Load(dt)

            if roomdata is None:
                continue

            for log in roomdata.Logs:

                id = log['ID']

                # 딕셔너리에 초기 데이터 기입
                if id not in times:
                    self.Statistics['_NAMES'][id] = log['Name']
                    self.Statistics[id] = [['', '', ''] for _ in range(len(self.FileDates))] 
                    times[id] = [[None, None] for _ in range(len(self.FileDates))] 

                # 입실 기록인 경우
                if log['Action'] == 'ChkIn':
                    ts = datetime.strptime(log['TimeStamp'], '%Y%m%d%H%M%S.%f')
                    if times[id][index][0] is None or times[id][index][0] > ts:
                        times[id][index][0] = ts

                # 퇴실 기록인 경우우
                elif log['Action'] == 'ChkOut':
                    ts = datetime.strptime(log['TimeStamp'], '%Y%m%d%H%M%S.%f')
                    if times[id][index][1] is None or times[id][index][1] < ts:
                        times[id][index][1] = ts


        # 취합한 시간 분석하고 통곗값 완성

        # ===============================
        # ⓒ 2025.05.05. GPT-4 generated
        # Logic: 교시별 출석 판단 알고리즘
        # Conditions: 시작/종료 시각과 허용 오차 고려
        # ===============================

        # 참고로 위에 쓰인 날짜, Logic, Conditions도 GPT-4가 써준거임.
        # GPT가 생성한 코드의 사용 범위를 물어보니까 자유롭게 써도 된다면서 주석 다는게 좋은 매너라고 저것도 써준거임.
        for id, timeLogs in times.items():
            for index, (in_time, out_time) in enumerate(timeLogs):

                if in_time is None and out_time is None:
                    self.Statistics[id][index] = ['', '', '']
                    continue

                for period in range(3):  # 3개 교시

                    # 기준 시간 정의
                    raw_date = self.FileDates[index]
                    start = datetime.combine(raw_date, RoomData.CONFIG.SelfStudyTimeData[period][2])
                    end = datetime.combine(raw_date, RoomData.CONFIG.SelfStudyTimeData[period][3])

                    # 자정 보정
                    if start.hour < 5:
                        start += timedelta(days=1)
                    if end.hour < 5:
                        end += timedelta(days=1)

                    # 허용 범위 보정
                    start_with_allowance = start + timedelta(minutes=RoomData.CONFIG.SelfStudyTimeData[period][4])
                    end_with_allowance = end - timedelta(minutes=RoomData.CONFIG.SelfStudyTimeData[period][4])

                    # 퇴실 None 처리
                    out_time_adj = out_time or end + timedelta(minutes=1)

                    # 기본적으로 X로 설정
                    status = 'X'

                    # 실제 입실/퇴실 시간이 해당 교시 시간과 겹치는지 확인
                    if in_time is not None and out_time_adj >= start and in_time <= end:
                        # 출입 시간의 일부라도 교시 시간 범위에 걸침

                        if in_time <= start_with_allowance and out_time_adj >= end_with_allowance:
                            status = 'O'
                        else:
                            status = '△'

                    self.Statistics[id][index][period] = status


    @staticmethod
    def format_timeStr(dtStr: str) -> str:
        if str(dtStr).strip().lower() in ('none', 'null', ''):
            return '?'

        dt = datetime.strptime(dtStr, '%Y%m%d%H%M%S.%f')

        # 오전/오후 결정
        period = "오전" if dt.hour < 12 else "오후"

        # 12시간제로 변환
        hour = dt.hour % 12
        if hour == 0:
            hour = 12

        # 초는 소수 첫째자리까지만
        seconds = f"{dt.second if dt.second > 9 else '0' + str(dt.second)}.{int(dt.microsecond / 100_000)}"

        # 최종 포맷
        return f"{period} {hour:02d}:{dt.minute:02d}:{seconds}"
    
    
    @staticmethod
    def _date(string: str) -> str:
        """ YYYYMMDD 형식의 날짜를 나타낸 문자열을 자르고 슬래시를 대입함. """
        return f'{string[0:4]}/{string[4:6]}/{string[6:8]}'


    @staticmethod
    def cell(sheet, row, column, value, setBold = False):
        sheet.cell(row=row, column=column).value = value
        sheet.cell(row=row, column=column).border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        sheet.cell(row=row, column=column).alignment = Alignment(
            vertical='center',
            horizontal='center'
        )
        sheet.cell(row=row, column=column).font = Font(bold=setBold)


    def Write(self) -> str:
        """ 통계를 작성하고 파일로 내보냄. 오류가 발생하면 기록을 출력하고 ErrorDialog를 표시함. """

        try:
            logging.info(f'교시별 출석부 통계 작성 시작: {self.Year}년 {self.Month}월')
            path = self._Write()
            logging.info(f'교시별 출석부 통계 작성 완료: {path}')
            return path
        except Exception as e:
            logging.error('교시별 출석부 통계 작성 중 오류가 발생하였습니다.', e, False, True)
            return None


    def _Write(self) -> str:

        wb: xl.Workbook = xl.load_workbook(RoomData.DIRECTORY + "/ChairyApp/sheets/period.xlsx")

        ws = wb.active

        ## 통계 작성
        index0 = 0

        # 상단 날짜 기록
        for index, dt in enumerate(self.Statistics['_DATA_DATES']):
            PeriodStatistics.cell(ws, 2, 4 + index * 3, PeriodStatistics._date(dt))

        # 교시별 기호 기록
        for index in range(len(self.Statistics['_DATA_DATES'])):
            PeriodStatistics.cell(ws, 3, 4 + index * 3, RoomData.CONFIG.SelfStudyTimeData[0][1], True)
            PeriodStatistics.cell(ws, 3, 5 + index * 3, RoomData.CONFIG.SelfStudyTimeData[1][1], True)
            PeriodStatistics.cell(ws, 3, 6 + index * 3, RoomData.CONFIG.SelfStudyTimeData[2][1], True)

        # 학생별 통계 기록
        for id, data in self.Statistics.items():

            if id in ('_DATA_DATES', '_NAMES'):
                continue

            index0 += 1

            row = 3 + index0

            # 순번, 학번, 이름 기록
            PeriodStatistics.cell(ws, row, 1, index0)
            PeriodStatistics.cell(ws, row, 2, id)

            nmCell = ws.cell(row=row, column=3)
            nmCell.value = self.Statistics['_NAMES'][id]
            nmCell.border = Border(
                                    left=Side(style='thin'),
                                    right=Side(style='thick'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin')
                                )
            nmCell.alignment = Alignment(
                                    vertical='center',
                                    horizontal='center'
                                )

            # 날짜별 통계 기록
            for index1, log in enumerate(data):

                # 통계 기록
                PeriodStatistics.cell(ws, row, 4 + index1 * 3, log[0])
                PeriodStatistics.cell(ws, row, 5 + index1 * 3, log[1])
                PeriodStatistics.cell(ws, row, 6 + index1 * 3, log[2])

        PeriodStatistics.cell(ws, 1, 8, f'{self.Year}년 {self.Month}월')

        for row in range(4, ws.max_row + 1):
            ws.row_dimensions[row].height = 21

        import os
        if not os.path.exists(RoomData.DIRECTORY + '/Statistics/'):
            os.makedirs(RoomData.DIRECTORY + '/Statistics/')

        finalFile = RoomData.DATA_DATE.strftime('교시별 출석부-%Y%m%d-') + datetime.now().strftime('%Y%m%d%H%M%S') + '.xlsx'
        wb.save(RoomData.DIRECTORY + '/Statistics/' + finalFile)
        
        return finalFile