import openpyxl as xl
from datetime import time



class ConfigurationError(Exception):

    def __init__(self, *args):
        super().__init__(*args)



class Configuration():
    """
    ### 구성 데이터

    '/school_data/configuration.xlsx'로부터 설정 값을 읽고 저장하는 클래스.
    """

    # Neis 연동 관련 설정
    NeisOpenApiKey  : str
    NeisOfficeCode  : str
    NeisSchoolCode  : str
    NeisCanIgnoreSSL: bool
    NeisTargetGrade : int

    # 휴업일, 공휴일 등의 쉬는 날에 지정석 적용 여부
    ReservedSeatInHoliday: bool

    # 학생 데이터
    Students    : list[list[str]]
    StudentIDs  : list[str]

    # 좌석 배치 데이터
    Arrangement : list[list]

    # 기타 소프트웨어 설정
    MediaDetection  : bool
    Alphabet        : bool

    # 자습 시간 정보
    SelfStudyTimeData: list[list]
    SelfStudyTimeVaild: bool



    def __init__(self, DIR: str):
        """ **DIR:** 'ChairyData._Dir'을 입력하면 됨. 실행 파일이 존재하는 디렉토리 위치 """
        self.Directory = DIR
        self.Init()


    def Init(self):

        wb = xl.load_workbook(self.Directory + "/school_data/configuration.xlsx", True, False, True, False, False)
        tmp: str


        ## 설정 및 구성 ##
        self.NeisOpenApiKey = wb['설정 및 구성']['B4'].value
        self.NeisOfficeCode = wb['설정 및 구성']['B5'].value
        self.NeisSchoolCode = wb['설정 및 구성']['B6'].value

        tmp = wb['설정 및 구성']['B7'].value
        if tmp == "O":
            self.NeisCanIgnoreSSL = True
        elif tmp == "X":
            self.NeisCanIgnoreSSL = False
        else:
            raise ConfigurationError("설정 및 구성 항목 중 'SSL 인증 무시 가능' 항목의 입력값이 잘못되었습니다. 입력 값: " + tmp)
        
        tmp = wb['설정 및 구성']['B8'].value
        try:
            self.NeisTargetGrade = int(tmp)
        except:
            raise ConfigurationError("설정 및 구성 항목 중 '대상 학년' 항목의 입력값 잘못되었습니다. 입력 값: " + tmp)
        
        tmp = wb['설정 및 구성']['B11'].value
        if tmp == "O":
            self.MediaDetection = True
        elif tmp == "X":
            self.MediaDetection = False
        else:
            raise ConfigurationError("설정 및 구성 항목 중 '\"지금 재생 중\" 기능' 항목의 입력값이 잘못되었습니다. 입력 값: " + tmp)
        
        tmp = wb['설정 및 구성']['B12'].value
        if tmp == "O":
            self.Alphabet = True
        elif tmp == "X":
            self.Alphabet = False
        else:
            raise ConfigurationError("설정 및 구성 항목 중 '알파벳 학번' 항목의 입력값이 잘못되었습니다. 입력 값: " + tmp)
        
        tmp = wb['설정 및 구성']['B13'].value
        if tmp == "O":
            self.ReservedSeatInHoliday = True
        elif tmp == "X":
            self.ReservedSeatInHoliday = False
        else:
            raise ConfigurationError("설정 및 구성 항목 중 '휴업일에도 지정석 시행' 항목의 입력값이 잘못되었습니다. 입력 값: " + tmp)
        

        ## 학생 정보 ##
        self.Students   = []
        self.StudentIDs = []
        verifyFailed = False
        try:
            tmp = ""
            for i in range(3, wb['학생 정보'].max_row + 1):
                
                # 좌측에 기입된 학번 불러오기
                tmp = str(wb['학생 정보'][f'A{i}'].value).strip()
                if tmp not in ("", "None"):

                    name = str(wb['학생 정보'][f'B{i}'].value).strip()
                    reservedSeat = str(wb['학생 정보'][f'C{i}'].value).strip()

                    # 기입 데이터 검증
                    try:
                        Configuration.VerifyStudentData(f'A{i}', tmp, f'B{i}', name)
                    except Exception as e:
                        verifyFailed = True
                        raise e

                    if reservedSeat not in ("", "None", "null"):
                        self.Students.append([tmp, name, reservedSeat])
                    else:
                        self.Students.append([tmp, name, None])
                    self.StudentIDs.append(tmp)

                # 우측에 기입된 학번 불러오기
                tmp = str(wb['학생 정보'][f'D{i}'].value).strip()
                if tmp not in ("", "None"):

                    name = str(wb['학생 정보'][f'E{i}'].value).strip()
                    reservedSeat = str(wb['학생 정보'][f'F{i}'].value).strip()

                    # 기입 데이터 검증
                    try:
                        Configuration.VerifyStudentData(f'D{i}', tmp, f'E{i}', name)
                    except Exception as e:
                        verifyFailed = True
                        raise e

                    if reservedSeat not in ("", "None", "null"):
                        self.Students.append([tmp, name, reservedSeat])
                    else:
                        self.Students.append([tmp, name, None])
                    self.StudentIDs.append(tmp)
        except Exception as e:
            if verifyFailed:
                raise e
            else:
                raise ConfigurationError("configuration.xlsx 파일의 '학생 정보' 시트를 읽는 도중 오류가 발생하였습니다. 파일을 확인해주십시오.")
        

        ## 자리 배치 ##
        self.Arrangement = []
        try:
            tmp = ""
            for i in range(2, wb['자리 배치'].max_row + 1):
                tmp = str(wb['자리 배치'][f'B{i}'].value).strip()
                if tmp in ["", "None"]:
                    continue
                
                self.Arrangement.append([tmp, int(wb['자리 배치'][f'C{i}'].value), int(wb['자리 배치'][f'D{i}'].value)])
        except:
            raise ConfigurationError("configuration.xlsx 파일의 '자리 배치' 시트를 읽는 도중 오류가 발생하였습니다. 파일을 확인해주십시오.")
        

        ## 자율학습 일정 정보 ##
        self.SelfStudyTimeData = [ [None, None, None, None, -1], [None, None, None, None, -1], [None, None, None, None, -1] ]
        
        for i in range(4, 7):
            self.SelfStudyTimeData[i - 4][0] = str(wb['설정 및 구성'][f'G{i}'].value).strip()
            self.SelfStudyTimeData[i - 4][1] = str(wb['설정 및 구성'][f'H{i}'].value).strip()
            self.SelfStudyTimeData[i - 4][2] = wb['설정 및 구성'][f'I{i}'].value
            self.SelfStudyTimeData[i - 4][3] = wb['설정 및 구성'][f'J{i}'].value
            try:
                self.SelfStudyTimeData[i - 4][4] = int(wb['설정 및 구성'][f'K{i}'].value)
            except:
                self.SelfStudyTimeVaild = False

        self.SelfStudyTimeVaild = True

        for i in range(0, 2):
            if self.SelfStudyTimeData[i][0] in ('None', ''):
                self.SelfStudyTimeVaild = False
                break
            if self.SelfStudyTimeData[i][1] in ('None', ''):
                self.SelfStudyTimeVaild = False
                break
            if not isinstance(self.SelfStudyTimeData[i][2], time):
                self.SelfStudyTimeVaild = False
                break
            if not isinstance(self.SelfStudyTimeData[i][3], time):
                self.SelfStudyTimeVaild = False
                break
        

        ## QR코드 ##
        """
        qr = str(wb['설정 및 구성']['B16'].value).strip()
        if qr not in ("", "None", "null"):
            self.QR_Active = True
            self.QR_Link = qr
            self.QR_Des1 = str(wb['설정 및 구성']['B17'].value).strip()
            self.QR_Des2 = str(wb['설정 및 구성']['B18'].value).strip()
            self.QR_Des3 = str(wb['설정 및 구성']['B19'].value).strip()
        else:
            self.QR_Active = False
        """
        

    @staticmethod
    def VerifyStudentData(id_cell: str, id: str, name_cell: str, name: str):
        """
        학번과 이름 형식이 적절한 지 확인함. 만약 문제가 발견되면 ConfigurationError를 일으킴.
        - - -
        #### 매개변수:
        - **id_cell:** 학번이 기입된 셀 위치
        - **id:** 기입된 학번
        - **name_cell:** 이름이 기입된 셀 위치
        - **name:** 기입된 이름
        """

        if len(id.strip()) != 4:
            raise ConfigurationError(f"구성 파일의 '학생 정보' 시트 중 [ {id_cell} ]번 셀의 학번 값은 네 자리가 입력되어야 합니다. 현재 {len(id.strip())} 자리가 입력되어 있습니다.")
        
        for letter in id.upper().strip():
            if letter not in ('_', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', '1', '2', '3', '4', '5', '6', '7', '8', '9', '0'):
                raise ConfigurationError(f"구성 파일의 '학생 정보' 시트 중 [ {id_cell} ]번 셀의 학번 값에는 알파벳과 숫자, '_'만 입력이 가능합니다. 현재 '{id.upper()}'이(가) 입력되어 있습니다.")
            
        if name.strip() == "None":
            raise ConfigurationError(f"구성 파일의 '학생 정보' 시트 중 [ {name_cell} ]번 셀의 이름 값에 'None'이 입력되어 있습니다. 해당 값은 시스템에서 내부적으로 사용되므로 입력이 허용되지 않습니다.")
        
        if name.strip() == "":
            raise ConfigurationError(f"구성 파일의 '학생 정보' 시트 중 [ {name_cell} ]번 셀의 이름 값이 입력되어 있지 않습니다.")