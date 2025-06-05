
from os import listdir
from os.path import exists
from datetime import date, datetime
from .RoomData import RoomData
import openpyxl as xl
from openpyxl.styles import *
from dataclasses import dataclass

from ..Logging import LoggingManager as logging



@dataclass(slots=True)
class MonthlyStatistics():
    """
    ### 월간 출석부 데이터

    특정한 달의 RoomData를 모두 취합하여 그 달의 일자별 입퇴실 시간, 이용 좌석을 통계로 만드는 클래스.
    """


    Year    : int
    Month   : int

    FileDates: list[date] # RoomData가 있는 날짜

    Statistics: dict
    # {
    #   '<StudentID>' : [ [<FirstChkIn>, <LastChkIn>, <LastSeat>], ... ]
    # }

    Empty: bool # 비어있음 여부



    def __init__(self, year: int, month: int):
        """
        #### 매개변수:
        - **year:** 내보내고자 하는 연도
        - **month:** 내보내고자 하는 달
        """

        # 날짜 설정
        self.Year  = year
        self.Month = month

        # 파일 이름 가져오고 날짜 추출
        if not exists(RoomData.DIRECTORY + date(year, month, 1).strftime('/RoomData/%Y%m/')):
            self.Empty = True
            return

        self.FileDates = []

        for fn in listdir(RoomData.DIRECTORY + date(year, month, 1).strftime('/RoomData/%Y%m/')):
            if fn.endswith('.json'):
                name = fn.split('.json')[0]
                try:
                    self.FileDates.append(datetime.strptime(name, '%Y%m%d').date())
                except Exception:
                    continue

        if len(self.FileDates) < 1:
            self.Empty = True
            return
        else:
            self.Empty = False

        # 시작 로그
        logging.info(f'월간 출석부 기록을 불러오는 중: {year}년 {month}월')

        # _DATA_DATES를 기본 포함해서 딕셔너리 구성
        self.Statistics = {
            '_DATA_DATES' : [dt.strftime('%Y%m%d') for dt in self.FileDates],
            '_NAMES' : {}
        }

        # 파일 별로 RoomData 불러오고 임시로 딕셔너리 작성
        for index, dt in enumerate(self.FileDates):

            roomdata = RoomData.Load(dt)

            if roomdata is None:
                continue

            for log in roomdata.Logs:
                self._Check(log['ID'], log['Name'])

                if log['Action'] == 'ChkIn':
                    self._ChkIn(index, log['ID'], datetime.strptime(log['TimeStamp'], '%Y%m%d%H%M%S.%f'), log['Seat'])
                elif log['Action'] == 'ChkOut':
                    self._ChkOut(index, log['ID'], datetime.strptime(log['TimeStamp'], '%Y%m%d%H%M%S.%f'))
                elif log['Action'] == 'Move':
                    self._LastSeat(index, log['ID'], log['Seat'])

        # 취합한 통계 모두 읽을 수 있는 문자열로 처리
        for id, data in self.Statistics.items():

            if id in ('_DATA_DATES', '_NAMES'):
                continue

            for index, log in enumerate(data):

                if log[0] is None and log[1] is None and log[2] is None:
                    self.Statistics[id][index] = ['', '', '']
                    continue

                if log[0] is not None:
                    self.Statistics[id][index][0] = MonthlyStatistics.format_time(log[0])
                else:
                    self.Statistics[id][index][0] = '-'

                if log[1] is not None:
                    self.Statistics[id][index][1] = MonthlyStatistics.format_time(log[1])
                else:
                    self.Statistics[id][index][1] = '-'

                if log[2] is None:
                    self.Statistics[id][index][2] = '-'


    def _Check(self, id: str, name: str):
        """
        딕셔너리에서 학번에 해당하는 데이터를 찾고, 없으면 새로 만들어냄.
        - - -
        #### 매개변수
        - **id:** 학번
        - **name:** 이름
        """
        if id not in self.Statistics:
            self.Statistics[id] = [[None, None, None] for _ in range(len(self.FileDates))] 
            self.Statistics['_NAMES'][id] = name
            
            
    def _ChkIn(self, index: int, id: str, firstChkIn: datetime, seat: str):
        """
        딕셔너리에서 특정 학번의 지정된 인덱스에 최초 입실 시간과 좌석을 기록함.

        만약 기존 데이터의 시간이 패러미터로 주어진 시간보다 더 이르면 변경하지 않음.
        - - -
        #### 매개변수:
        - **index:** 순번(또는 인덱스)
        - **id:** 학번
        - **firstChkIn:** 입실 시간
        - **seat:** 입실 좌석
        """
        if self.Statistics[id][index][0] is None or self.Statistics[id][index][0] > firstChkIn:
            self.Statistics[id][index][0] = firstChkIn
            self.Statistics[id][index][2] = seat


    def _ChkOut(self, index: int, id: str, lastChkOut: datetime):
        """
        딕셔너리에서 특정 학번의 지정된 인덱스에 마지막 퇴실 시간을 기록함.

        만약 기존 데이터의 시간이 패러미터로 주어진 시간보다 더 늦으면 변경하지 않음.
        - - -
        #### 매개변수:
        - **index:** 순번(또는 인덱스)
        - **id:** 학번
        - **firstChkIn:** 퇴실 시간
        """
        if self.Statistics[id][index][1] is None or self.Statistics[id][index][1] < lastChkOut:
            self.Statistics[id][index][1] = lastChkOut


    def _LastSeat(self, index: int, id: str, lastSeat: int):
        """
        딕셔너리에서 특정 학번의 지정된 인덱스에 마지막 이용 좌석을 기록함.
        - - -
        #### 매개변수:
        - **index:** 순번(또는 인덱스)
        - **id:** 학번
        - **lastSeat:** 최종 이용 좌석
        """
        self.Statistics[id][index][2] = lastSeat


    @staticmethod
    def format_timeStr(dtStr: str) -> str:
        if str(dtStr).strip().lower() in ('none', 'null', ''):
            return '?'

        dt = datetime.strptime(dtStr, '%Y%m%d%H%M%S.%f')

        # 오전/오후 결정
        period = "오전" if dt.hour < 12 else "오후"

        # 12시간제로 변환
        hour = dt.hour % 12
        if hour == 0:
            hour = 12

        # 초는 소수 첫째자리까지만
        seconds = f"{dt.second if dt.second > 9 else '0' + str(dt.second)}.{int(dt.microsecond / 100_000)}"

        # 최종 포맷
        return f"{period} {hour:02d}:{dt.minute:02d}:{seconds}"
    
    
    @staticmethod
    def _date(string: str) -> str:
        return f'{string[0:4]}년 {string[4:6]}월 {string[6:8]}일'


    @staticmethod
    def cell(sheet, row, column, value):
        sheet.cell(row=row, column=column).value = value
        sheet.cell(row=row, column=column).border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        sheet.cell(row=row, column=column).alignment = Alignment(
            vertical='center',
            horizontal='center'
        )
        #sheet.cell(row=row, column=column).font = Font(bold=True)


    def Write(self) -> str:
        """ 통계를 작성하고 파일로 내보냄. 오류가 발생하면 기록을 출력하고 ErrorDialog를 표시함. """

        try:
            logging.info(f'월간 출석부 통계 작성 시작: {self.Year}년 {self.Month}월')
            path = self._Write()
            logging.info(f'월간 출석부 통계 작성 완료: {path}')
            return path
        except Exception as e:
            logging.error('월간 출석부 통계 작성 중 오류가 발생하였습니다.', e, False, True)
            return None


    def _Write(self) -> str:

        wb: xl.Workbook = xl.load_workbook(RoomData.DIRECTORY + "/ChairyApp/sheets/monthly.xlsx")

        ws = wb.active

        ## 통계 작성
        index0 = 0

        # 상단 날짜 기록
        for index, dt in enumerate(self.Statistics['_DATA_DATES']):
            MonthlyStatistics.cell(ws, 2, 4 + index * 3, MonthlyStatistics._date(dt))

        # 학생별 통계 기록
        for id, data in self.Statistics.items():

            if id in ('_DATA_DATES', '_NAMES'):
                continue

            index0 += 1

            row = 3 + index0

            # 순번, 학번, 이름 기록
            MonthlyStatistics.cell(ws, row, 1, index0)
            MonthlyStatistics.cell(ws, row, 2, id)

            nmCell = ws.cell(row=row, column=3)
            nmCell.value = self.Statistics['_NAMES'][id]
            nmCell.border = Border(
                                    left=Side(style='thin'),
                                    right=Side(style='thick'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin')
                                )
            nmCell.alignment = Alignment(
                                    vertical='center',
                                    horizontal='center'
                                )

            # 날짜별 통계 기록
            for index1, log in enumerate(data):

                # 통계 기록
                MonthlyStatistics.cell(ws, row, 4 + index1 * 3, log[0])
                MonthlyStatistics.cell(ws, row, 5 + index1 * 3, log[1])
                MonthlyStatistics.cell(ws, row, 6 + index1 * 3, log[2])

        MonthlyStatistics.cell(ws, 1, 8, f'{self.Year}년 {self.Month}월')

        for row in range(4, ws.max_row + 1):
            ws.row_dimensions[row].height = 21

        import os
        if not os.path.exists(RoomData.DIRECTORY + '/Statistics/'):
            os.makedirs(RoomData.DIRECTORY + '/Statistics/')

        finalFile = RoomData.DATA_DATE.strftime('월간 출석부-%Y%m%d-') + datetime.now().strftime('%Y%m%d%H%M%S') + '.xlsx'
        wb.save(RoomData.DIRECTORY + '/Statistics/' + finalFile)
        
        return finalFile



    @staticmethod
    def format_time(dt: datetime) -> str:

        if dt == None:
            return '?'
        
        # 오전/오후 결정
        period = "오전" if dt.hour < 12 else "오후"

        # 12시간제로 변환
        hour = dt.hour % 12
        if hour == 0:
            hour = 12

        # 초는 소수 첫째자리까지만
        seconds = f"{dt.second if dt.second > 9 else '0' + str(dt.second)}.{int(dt.microsecond / 100_000)}"

        # 최종 포맷
        return f"{period} {hour:02d}:{dt.minute:02d}:{seconds}"